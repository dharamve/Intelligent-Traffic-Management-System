import random
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

RED_SIGNAL_INTERVAL = 15  
RED_SIGNAL_DURATION = 3  
MAX_SPEED = 100

car1_speed = random.randint(50, MAX_SPEED)
car2_speed = random.randint(50, MAX_SPEED)
car3_speed = random.randint(50, MAX_SPEED)

lowest_speed = min(car1_speed, car2_speed, car3_speed)

wb = Workbook()
ws = wb.active

ws.append(["Date", "Time", "Car 1 Speed", "Car 2 Speed", "Car 3 Speed", "Status", "Average Speed"])

wb.save("traffic_data.xlsx")

red_signal_counter = 0
is_red_signal = False
red_signal_remaining_duration = 0

total_speed = 0
num_iterations = 0

iteration_nums = []
car1_speeds = []
car2_speeds = []
car3_speeds = []
average_speeds = []

while True:
    now = datetime.now()

    current_date = now.strftime("%d/%m/%Y")
    current_time = now.strftime("%I:%M:%S %p")

    if red_signal_counter % RED_SIGNAL_INTERVAL == 0:
        is_red_signal = True
        red_signal_counter = 1  
        red_signal_remaining_duration = RED_SIGNAL_DURATION
    else:
        is_red_signal = False
        red_signal_counter += 1

    if is_red_signal and red_signal_remaining_duration > 0:
        car1_speed = 0
        car2_speed = 0
        car3_speed = 0
        red_signal_remaining_duration -= 1
        status = "Red Signal"
    else:
        if red_signal_remaining_duration <= -3:
            car1_speed += (lowest_speed - 40) // (RED_SIGNAL_DURATION - 2)
            car2_speed += (lowest_speed - 40) // (RED_SIGNAL_DURATION - 2)
            car3_speed += (lowest_speed - 40) // (RED_SIGNAL_DURATION - 2)
        elif red_signal_remaining_duration == -2:
            car1_speed += (lowest_speed - 40) // (RED_SIGNAL_DURATION - 1)
            car2_speed += (lowest_speed - 40) // (RED_SIGNAL_DURATION - 1)
            car3_speed += (lowest_speed - 40) // (RED_SIGNAL_DURATION - 1)
        elif red_signal_remaining_duration == -1:
            car1_speed += (lowest_speed - 40) // RED_SIGNAL_DURATION
            car2_speed += (lowest_speed - 40) // RED_SIGNAL_DURATION
            car3_speed += (lowest_speed - 40) // RED_SIGNAL_DURATION
        else:
            car1_speed = max(car1_speed + random.randint(-2, 2), lowest_speed - 40)
            car2_speed = max(car2_speed + random.randint(-2, 2), lowest_speed - 40)
            car3_speed = max(car3_speed + random.randint(-2, 2), lowest_speed - 40)

        status = "Green Signal"

    print("Car 1 Speed:", car1_speed)
    print("Car 2 Speed:", car2_speed)
    print("Car 3 Speed:", car3_speed)

    total_speed += car1_speed + car2_speed + car3_speed
    num_iterations += 1

    average_speed = round(total_speed / (num_iterations*2.5), 2)

    wb = load_workbook("traffic_data.xlsx")
    ws = wb.active

    ws.append([current_date, current_time, car1_speed, car2_speed, car3_speed, status, average_speed])

    wb.save("traffic_data.xlsx")

    iteration_nums.append(num_iterations)
    car1_speeds.append(car1_speed)
    car2_speeds.append(car2_speed)
    car3_speeds.append(car3_speed)
    average_speeds.append(average_speed)

    if num_iterations % 25 == 0:
        # Plot line graph
        plt.plot(iteration_nums, car1_speeds, label="Car 1 Speed")
        plt.plot(iteration_nums, car2_speeds, label="Car 2 Speed")
        plt.plot(iteration_nums, car3_speeds, label="Car 3 Speed")
        plt.plot(iteration_nums, average_speeds, label="Average Speed")
        plt.xlabel("Time(in seconds)")
        plt.ylabel("Speed")
        plt.title("Speed Variation")
        plt.legend()
        plt.grid(True)
        plt.show()

    if car1_speed == 0 and car2_speed == 0 and car3_speed == 0:
        status = "Red Signal"

    time.sleep(1)
